from flask import request,render_template
import random
import openpyxl
from werkzeug.security import generate_password_hash, check_password_hash
from exam_portal_end_to_end.api.config import app
from exam_portal_end_to_end.api.model import *

import base64

FILE_PATH = 'C:/Users/Ravi/PycharmProjects/Flask/exam_portal_end_to_end/api/files/'

def deserialize_data(data):
    questionlist = []
    for que in data:
        questionlist.append({'Qid':que.id,'Question': que.question,'Qop1':que.option1,'Qop2':que.option2,
                             'Qop3':que.option3,
                             'Qop4':que.option4,'Qans':que.answer})
    return questionlist

def file_encode():
    with open(FILE_PATH + 'Questions.xlsx', 'rb') as f:
        file = f.read()
        binary_code = base64.b64encode(file)
        binary_msg = binary_code.decode('utf-8')
        return binary_msg


@app.route('/client/registration/', methods=['POST'])
def client_registration():
    reqdata = request.get_json()
    if reqdata:
        if reqdata.get('fullname') and reqdata.get('email') and reqdata.get('username') and reqdata.get('password'):
            dbclient = Client.query.filter(Client.username == reqdata.get('username')).first()
            if not dbclient:
                client = Client(fullname=reqdata.get('fullname'),
                                email=reqdata.get('email'),
                                username=reqdata.get('username'),
                                password=generate_password_hash(reqdata.get('password')))
                db.session.add(client)
                db.session.commit()
                return {'success': 'Client Registration Successfully..!'}
            return {'error': 'Duplicate username..!'}
    return {'error': 'All fields required [fullname,email,username,password]..!'}


@app.route('/student/registration/', methods=['POST'])
def student_registration():
    reqdata = request.get_json()
    if reqdata:
        if reqdata.get('fullname') and reqdata.get('email') and reqdata.get('username') and reqdata.get(
                'password') and reqdata.get('employer id'):
            dbstud = Student.query.filter(Student.username == reqdata.get('username')).first()
            if not dbstud:
                studexaminer = Client.query.filter_by(id=reqdata.get('employer id')).first()
                if studexaminer:
                    stud = Student(fullname=reqdata.get('fullname'),
                                   email=reqdata.get('email'),
                                   username=reqdata.get('username'),
                                   password=generate_password_hash(reqdata.get('password')),
                                   clientid=reqdata.get('employer id'))
                    db.session.add(stud)
                    db.session.commit()
                    return {'success': 'Student Registration Successfully..!'}
                return {'error': 'Examiner Given id not present..!'}
            return {'error': 'Duplicate username..!'}
    return {'error': 'All fields required [fullname,email,username,password,employer id]..!'}


@app.route('/client/token/', methods=['PATCH'])
def create_client_token():
    reqdata = request.get_json()
    if reqdata:
        if reqdata.get('username') and reqdata.get('password'):
            dbclient = Client.query.filter(Client.username == reqdata.get('username')).first()
            if dbclient and check_password_hash(dbclient.password, reqdata.get('password')):
                dbclient.token = generate_password_hash(
                    dbclient.username + dbclient.password + str(random.randint(1111, 9999)))
                db.session.commit()
                return {'success': dbclient.token}
            return {'error': 'Invalid username and password..!'}
    return {'error': 'All fields required [username,password]..!'}


@app.route('/student/login/', methods=['POST'])
def student_login():
    reqdata = request.get_json()
    if reqdata:
        if reqdata.get('username') and reqdata.get('password'):
            dbstud = Student.query.filter(Student.username == reqdata.get('username')).first()
            if dbstud and check_password_hash(dbstud.password, reqdata.get('password')):
                questions = dbstud.clref.questions
                return {'success': dbstud.clientid}
            return {'error': 'Invalid username and password..!'}
    return {'error': 'All fields required [username,password]..!'}


@app.route('/client/login/', methods=['POST'])
def client_login():
    reqdata = request.get_json()
    if reqdata:
        if reqdata.get('username') and reqdata.get('password'):
            dbclient = Client.query.filter(Client.username == reqdata.get('username')).first()
            if dbclient and check_password_hash(dbclient.password, reqdata.get('password')):
                questions = dbclient.questions
                if dbclient.token:
                    return {'token': dbclient.token, 'file': file_encode()}
                return {'error': 'Unauthorized access generate token ..!'}
            return {'error': 'Invalid username and password..!'}
    return {'error': 'All fields required [username,password]..!'}


@app.route('/student/question/', methods=['GET'])
def student_question():
    response = request.get_json()
    id = response.get('clientid')
    if id:
        client = Client.query.filter_by(id=id).first()
        if client:
            return {'success':deserialize_data(client.questions)}
        return {'error': 'Not present..!'}
    return {'error': 'All fields required [username,password]..!'}

@app.route("/upload/questions/",methods=['POST'])
def read_excelsheet():
    reqdata = request.files
    if reqdata:
        f = request.files['file']
        f.save(FILE_PATH+request.form.get('name')+'.xlsx')
        errors = {}
        dbclient = Client.query.filter(Client.username==request.form.get('name')).first()
        try:
            workbook = openpyxl.load_workbook(FILE_PATH+request.form.get('name')+'.xlsx')
            sheet = workbook['questions']  # what if sheet name not present
        except BaseException as e:
            errors = {"file": "{}".format(e.args)}
            print(e.args)
        else:
            max_row = sheet.max_row
            ques_list = []
            for row in range(1, max_row + 1):
                if row == 1:  # header
                    continue
                a1 = sheet.cell(row, 1).value
                a2 = sheet.cell(row, 2).value
                a3 = sheet.cell(row, 3).value
                a4 = sheet.cell(row, 4).value
                a5 = sheet.cell(row, 5).value
                a6 = sheet.cell(row, 6).value  # if no 6 columns

                values = {a2, a3, a4, a5}
                if len(values) == 4 and a6 in values:
                    ques_list.append(Questions(question=a1, option1=a2, option2=a3, option3=a4, option4=a5,
                                               answer=a6,
                                               clientid=dbclient.id))
                else:
                    errors[row] = 'Not upload in database..'
            if not errors:  # no errors -- tr ch question upload hotil
                db.session.add_all(ques_list)
                db.session.commit()
                return {'success': 'File Upload successfully..!'}
        return {'error': errors}
    return {'error': 'no data send'}

@app.route('/user/questions/')
def client_questions():
    clnt = Client.query.all()
    return render_template('client_questions.html',clntlist=clnt)


if __name__ == '__main__':
    db.create_all()
    app.run(debug=True)
